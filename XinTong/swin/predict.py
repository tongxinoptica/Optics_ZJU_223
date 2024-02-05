import torch
from torchvision import transforms
from swin_IR import TWC_Swin
# from unet_model import Unet
# from model import RDR_model
from unit import to_psnr, to_ssim_skimage, to_ssim, to_pearson
from PIL import Image
import matplotlib.pyplot as plt
import torch.nn as nn
import openpyxl

# wb = openpyxl.Workbook()
# ws = wb.create_sheet('celeb')  # 新建表格名称
# ws.cell(row=2, column=1).value = "mse"
# ws.cell(row=3, column=1).value = "psnr"
# ws.cell(row=4, column=1).value = "npcc"
# ws.cell(row=5, column=1).value = "ssim"
# s = 1

# Parameter setting
coherence = '1.1f'
name = ['4熊', '339飞机', '666人', '2893风景', '5642人', '5864碗', '6716叶子', '8544番茄']
num = '666'
Image_path = 'C:/Users/Administrator/Desktop/paper3/tur_image/ocean/0.1/jpg/{}_in.jpg'.format(num)
label_path = 'C:/Users/Administrator/Desktop/paper3/tur_image/jpg/{}.jpg'.format(num)

Tensor = transforms.Compose([transforms.Resize((128, 128)), transforms.ToTensor()])
model = TWC_Swin()
device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")  # 定义训练的设备
model.to(device)
print("Model loaded")


img = Image.open(Image_path)
if img.mode=='RGB':
    img = img.convert('L')
img = Tensor(img)
img = img / (torch.max(img))  # 归一化0-1
img = torch.reshape(img, (1, 1, 128, 128))

label = Image.open(label_path)
if label.mode=='RGB':
    label = label.convert('L')
label = Tensor(label)
label = label / (torch.max(label))  # 归一化0-1
label = torch.reshape(label, (1, 1, 128, 128))  # 增加batch维度

model.load_state_dict(torch.load('./parameter/lr=0.0005_{}_nor'.format(coherence)))
#model.load_state_dict(torch.load('./parameter/swin_noconv_{}'.format(parameter)))
b = torch.ones(128, 128).to(device)

model.eval()
with torch.no_grad():
    img = img.to(device)
    label = label.to(device)
    output = model(img)
    output = torch.where(output < 1, output, b)
    # print(torch.max(output))
    mseloss = nn.MSELoss()
    mse = mseloss(output, label)
    psnr = to_psnr(output, label)
    npcc = to_pearson(output, label)
    print(mse)
    print(psnr)
    ssim = to_ssim_skimage(output, label)
    ssim1 = to_ssim(output, label)
    print(ssim)
    print(npcc)
    # print(ssim1)
    # s+=1
    # ws.cell(row=1, column=s).value = num
    # ws.cell(row=2, column=s).value = mse.item()
    # ws.cell(row=3, column=s).value = psnr
    # ws.cell(row=4, column=s).value = ssim[0]
    # ws.cell(row=5, column=s).value = npcc.item()
    to_pil = transforms.ToPILImage()
    output_pil = to_pil(output[0])  # 删去batch维度
    gt = to_pil(label[0])
    plt.imshow(output_pil, cmap='gray')
    plt.axis('off')
    #plt.savefig('C:/Users/Administrator/Desktop/paper3/tur_image/compare/our/{}_out_o0.1.jpg'.format(num),bbox_inches='tight', pad_inches=0)
    plt.show()


