import os
import torch
from torch import nn
from torchvision import transforms
from swin_IR import swinIR
from PIL import Image
from unit import to_psnr, to_ssim_skimage, to_ssim, to_pearson
import tqdm
import openpyxl
import time
# Change Parameter
tur_image = '8544番茄'
tur_image_gt = '8544.jpg'
tur_intensity = '50'
parameter = ['f','1.1f','1.2f','1.3f','1.4f','1.5f']
name = 'WED'
num = ['4','666','5864','6716','8544']
s = 1
print('Image number: 100')

wb = openpyxl.Workbook()
ws = wb.create_sheet('celeb')  # 新建表格名称
ws.cell(row=2, column=1).value = "mse"
ws.cell(row=3, column=1).value = "psnr"
ws.cell(row=4, column=1).value = "npcc"
ws.cell(row=5, column=1).value = "ssim"

for p in parameter:
    Tensor = transforms.Compose([transforms.Resize((128, 128)), transforms.ToTensor()])
    model = swinIR()
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")  # 定义训练的设备
    model.to(device)
    model.load_state_dict(torch.load('./parameter/lr=0.0005_{}_nor'.format(p)))  # 加载权重
    #model.load_state_dict(torch.load('./parameter/swin_noconv_{}'.format(p)))
    print("Swin-model loading success, Model space: {}".format(p))
    # img_path = 'D:/turbulence/tur_image/air/{}/{}'.format(tur_intensity,tur_image)
    # lb_path = 'D:/HR_data/exp_data/train_gt/{}'.format(tur_image_gt)
    img_path = 'D:/HR_data/exp_data/{}/test/{}'.format(p, name)
    lb_path = 'D:/HR_data/exp_data/test_gt/{}'.format(name)
    imglist = os.listdir(img_path)
    imglist.sort(key=lambda x: int(x[:-4]))  # 倒着数第四位'.'为分界线，按照‘.'左边的数字从小到大排序
    #s = 1
    total_mse = 0
    total_psnr = 0
    total_ssim = 0
    total_npcc = 0
    start = time.time()
    for i in imglist:
        image_path = os.path.join(img_path, i)
        label_path = os.path.join(lb_path, i)
        image = Image.open(image_path)
        image = Tensor(image)
        image = image / (torch.max(image))  # 归一化0-1
        image = torch.reshape(image, (1, 1, 128, 128))  # 增加batch维度
        # label = Image.open(label_path)
        label = Image.open(label_path)
        label = Tensor(label)
        #label = label / (torch.max(label))  # 归一化0-1
        label = torch.reshape(label, (1, 1, 128, 128))
        b = torch.ones(128, 128).to(device)

        model.eval()
        with torch.no_grad():
            img = image.to(device)
            label = label.to(device)
            output = model(img)
            output = torch.where(output < 1, output, b)
            #print(torch.max(output))
            mse_loss = nn.MSELoss()
            mse = mse_loss(output, label)
            # print(mse)
            psnr = to_psnr(output, label)
            npcc = to_pearson(output, label)
            # print(psnr)
            # print(npcc)
            # ssim = to_ssim_skimage(output, label)
            ssim = to_ssim_skimage(output, label)
            # print(ssim)
            # print(ssim1)
            total_mse += mse
            total_psnr += psnr
            total_ssim += ssim[0]
            total_npcc += npcc
    end = time.time()
    print('The running time is {} s'.format(end-start))
    mse = total_mse/100
    psnr = total_psnr/100
    ssim = total_ssim/100
    npcc = total_npcc/100
    s += 1
    ws.cell(row=1, column=s).value = p
    ws.cell(row=2, column=s).value = mse.item()
    ws.cell(row=3, column=s).value = psnr
    ws.cell(row=4, column=s).value = ssim
    ws.cell(row=5, column=s).value = npcc.item()
#wb.save('1.xlsx')

