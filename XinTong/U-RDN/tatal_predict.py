import os
import torch
from torch import nn
from torchvision import transforms
from model import RDR_model
from PIL import Image
from unit import to_psnr, to_ssim_skimage, to_ssim, to_pearson
import openpyxl
import time

wb = openpyxl.Workbook()
ws = wb.create_sheet('voc')  # 新建表格名称
ws.cell(row=1, column=1).value = "mse"
ws.cell(row=1, column=2).value = "psnr"
ws.cell(row=1, column=3).value = "npcc"
ws.cell(row=1, column=4).value = "ssim"

Tensor = transforms.Compose([transforms.Resize((256, 256)), transforms.ToTensor()])
RDR_model = RDR_model()
device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")  # 定义训练的设备
RDR_model.to(device)
RDR_model.load_state_dict(torch.load('./parameter/mse_gr36_f'))  # 加载权重
print("模型加载完成")
img_path = 'D:/paper2_total_data/test/d=f/input_data/voc'
lb_path = 'D:/paper2_total_data/test/d=f/groundtruth_data/voc'

imglist = os.listdir(img_path)
imglist.sort(key=lambda x: int(x[:-4]))  # 倒着数第四位'.'为分界线，按照‘.'左边的数字从小到大排序
s = 1
start = time.time()
for i in imglist:
    image_path = os.path.join(img_path, i)
    label_path = os.path.join(lb_path, i)
    image = Image.open(image_path)
    image = Tensor(image)
    image = torch.reshape(image, (1, 1, 256, 256))  # 增加batch维度
    label = Image.open(label_path)
    label = Tensor(label)
    label = torch.reshape(label, (1, 1, 256, 256))
    b = torch.ones(256, 256).to(device)

    RDR_model.eval()

    with torch.no_grad():
        img = image.to(device)
        label = label.to(device)
        output = RDR_model(img)
        output = torch.where(output < 1, output, b)
        # print(torch.max(output))
        mse_loss = nn.MSELoss()
        mse = mse_loss(output, label)
        # print(mse)
        psnr = to_psnr(output, label)
        npcc = to_pearson(output, label)
        # print(psnr)
        # print(npcc)
        #ssim = to_ssim_skimage(output, label)
        ssim = to_ssim(output, label)
        # print(ssim)
        # print(ssim1)
end = time.time()
print(start-end)
        # s += 1
        # ws.cell(row=s, column=1).value = mse.item()
        # ws.cell(row=s, column=2).value = psnr
        # ws.cell(row=s, column=3).value = npcc.item()
        # ws.cell(row=s, column=4).value = ssim.item()
        # wb.save('D:/test/npcc_voc.xlsx')
