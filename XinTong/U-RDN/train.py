import torch
import torch.nn as nn
import time
import openpyxl
from torch.cuda import empty_cache
import torch.nn.functional as F
from train_dataloader import train_data
from val_dataloader import val_data
from torch.utils.data import DataLoader
from model import RDR_model
from unit import to_psnr, to_ssim_skimage, to_ssim, to_mseloss, to_pearson
from NPCC import NPCC_loss, ssim_loss, mse_loss


# 创建表格
# wb = openpyxl.Workbook()
# ws = wb.create_sheet('epoch_d=f')  # 新建表格名称
# 加载数据
# imagenet数据集
train_root_dir = 'D:/train/d=1.5f/input_data/imagenet'
train_label_dir = 'D:/train/d=1.5f/groundtruth_data/imagenet'
train_loader = DataLoader(train_data(train_root_dir, train_label_dir), batch_size=8, shuffle=True)
# voc验证集
voc_val_root_dir = 'D:/val/d=1.5f/input_data/voc'
voc_val_label_dir = 'D:/val/d=1.5f/groundtruth_data/voc'
voc_val_loader = DataLoader(val_data(voc_val_root_dir, voc_val_label_dir), batch_size=25)
# coco验证集
coco_val_root_dir = 'D:/val/d=1.5f/input_data/coco'
coco_val_label_dir = 'D:/val/d=1.5f/groundtruth_data/coco'
coco_val_loader = DataLoader(val_data(coco_val_root_dir, coco_val_label_dir), batch_size=25)
# lfw验证集
lfw_val_root_dir = 'D:/val/d=1.5f/input_data/lfw'
lfw_val_label_dir = 'D:/val/d=1.5f/groundtruth_data/lfw'
lfw_val_loader = DataLoader(val_data(lfw_val_root_dir, lfw_val_label_dir), batch_size=25)
# celeb验证集
celeb_val_root_dir = 'D:/val/d=1.5f/input_data/celeb'
celeb_val_label_dir = 'D:/val/d=1.5f/groundtruth_data/celeb'
celeb_val_loader = DataLoader(val_data(celeb_val_root_dir, celeb_val_label_dir), batch_size=25)
# celeb验证集
img_val_root_dir = 'D:/val/d=1.5f/input_data/imagenet'
img_val_label_dir = 'D:/val/d=1.5f/groundtruth_data/imagenet'
img_val_loader = DataLoader(val_data(img_val_root_dir, img_val_label_dir), batch_size=25)

# 加载模型
RDR_model = RDR_model()
device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")  # 定义训练的设备
RDR_model.to(device)
print("模型加载完成，开始训练")

# 损失函数

loss_function = nn.MSELoss()
loss_function3 = NPCC_loss()
# loss_function = ssim_loss()
loss_function1 = mse_loss()
loss_function2 = nn.SmoothL1Loss()
# loss_function.to(device)

learning_rate = 0.002

# weight_decay_list = (param for name, param in RDR_model.named_parameters() if name[-4:] != 'bias' and "bn" not in name)
# no_decay_list = (param for name, param in RDR_model.named_parameters() if name[-4:] == 'bias' or "bn" in name)
# parameters = [{'params': weight_decay_list},
#               {'params': no_decay_list, 'weight_decay': 0.}]

optimizer = torch.optim.Adam(RDR_model.parameters(), lr=learning_rate)

# 加载权重
# try:
#     RDR_model.load_state_dict(torch.load('mse_gr36_1.5f'))
#     print('--- weight loaded ---')
# except:
#     print('--- no weight loaded ---')

total_epoch = 120
old_psnr = 0.0
old_ssim = 0.0
accumulation_steps = 4
# b = torch.ones(256, 256).to(device)

s = 1
for i in range(total_epoch):
    print("epoch:{}".format(i + 1))
    start_time = time.time()
    if ((i + 1) % 10) == 0:
        learning_rate /= 2
    # 后期加入学习率调整
    # 训练
    # empty_cache()
    RDR_model.train()
    psnr_list = []
    ssim_list = []
    mse_list = []
    npcc_list = []
    for batch_id, train_data in enumerate(train_loader):
        imgs, label = train_data
        imgs = imgs.to(device)
        label = label.to(device)
        output = RDR_model(imgs)  # 计算输出
        # output = torch.where(output < 1, output, b)

        #loss = loss_function(output, label)  # 计算损失函数
        loss = loss_function1(output, label) + loss_function3(output, label)
        # print(loss)
        # optimizer.zero_grad()
        loss /= accumulation_steps
        loss.backward()
        # optimizer.step()
        if ((batch_id + 1) % accumulation_steps) == 0:  # 梯度叠加
            optimizer.step()
            optimizer.zero_grad()

            # Calculate train average psnr, ssim
            psnr_list.append(to_psnr(output, label))
            ssim_list.append(to_ssim(output, label))
            mse_list.append(to_mseloss(output, label))
            npcc_list.append(to_pearson(output, label))

    train_psnr = sum(psnr_list) / len(psnr_list)
    train_ssim = sum(ssim_list) / len(ssim_list)
    train_mse = sum(mse_list) / len(mse_list)
    train_npcc = sum(npcc_list) / len(npcc_list)
    print("Average train mse:{}".format(train_mse))
    print("Average train psnr:{}".format(train_psnr))
    print("Average train ssim:{}".format(train_ssim))
    print("Average train npcc:{}".format(train_npcc))
    s += 1
    # ws.cell(row=1, column=1).value = "train_mse"
    # ws.cell(row=1, column=2).value = "train_psnr"
    # ws.cell(row=1, column=3).value = "train_npcc"
    # ws.cell(row=1, column=4).value = "train_ssim"
    # ws.cell(row=s, column=1).value = train_mse.item()
    # ws.cell(row=s, column=2).value = train_psnr
    # ws.cell(row=s, column=3).value = train_npcc.item()
    # ws.cell(row=s, column=4).value = train_ssim.item()

    RDR_model.eval()
    # 验证voc数据集
    voc_psnr_list_ = []
    voc_ssim_list_ = []
    voc_mse_list_ = []
    voc_npcc_list_ = []

    for batch_id, val_data in enumerate(voc_val_loader):
        with torch.no_grad():
            imgs, label = val_data
            imgs = imgs.to(device)
            label = label.to(device)
            output = RDR_model(imgs)
            # output = torch.where(output < 1, output, b)

        # Calculate  eval average psnr, ssim

        voc_psnr_list_.append(to_psnr(output, label))
        voc_ssim_list_.append(to_ssim(output, label))
        voc_mse_list_.append(to_mseloss(output, label))
        voc_npcc_list_.append(to_pearson(output, label))

    voc_val_psnr = sum(voc_psnr_list_) / len(voc_psnr_list_)
    voc_val_ssim = sum(voc_ssim_list_) / len(voc_ssim_list_)
    voc_val_mse = sum(voc_mse_list_) / len(voc_mse_list_)
    voc_val_npcc = sum(voc_npcc_list_) / len(voc_npcc_list_)
    print("Average eval voc_psnr:{},voc_ssim:{}".format(voc_val_psnr, voc_val_ssim))
    print("Average eval voc_mse:{},voc_npcc:{}".format(voc_val_mse, voc_val_npcc))
    # ws.cell(row=1, column=5).value = "voc_mse"
    # ws.cell(row=1, column=6).value = "voc_psnr"
    # ws.cell(row=1, column=7).value = "voc_npcc"
    # ws.cell(row=1, column=8).value = "voc_ssim"
    # ws.cell(row=s, column=5).value = voc_val_mse.item()
    # ws.cell(row=s, column=6).value = voc_val_psnr
    # ws.cell(row=s, column=7).value = voc_val_npcc.item()
    # ws.cell(row=s, column=8).value = voc_val_ssim.item()

    # 验证coco数据集
    coco_psnr_list_ = []
    coco_ssim_list_ = []
    coco_mse_list_ = []
    coco_npcc_list_ = []
    for batch_id, val_data in enumerate(coco_val_loader):
        with torch.no_grad():
            imgs, label = val_data
            imgs = imgs.to(device)
            label = label.to(device)
            output = RDR_model(imgs)
            # output = torch.where(output < 1, output, b)

        # Calculate  eval average psnr, ssim

        coco_psnr_list_.append(to_psnr(output, label))
        coco_ssim_list_.append(to_ssim(output, label))
        coco_mse_list_.append(to_mseloss(output, label))
        coco_npcc_list_.append(to_pearson(output, label))

    coco_val_psnr = sum(coco_psnr_list_) / len(coco_psnr_list_)
    coco_val_ssim = sum(coco_ssim_list_) / len(coco_ssim_list_)
    coco_val_mse = sum(coco_mse_list_) / len(coco_mse_list_)
    coco_val_npcc = sum(coco_npcc_list_) / len(coco_npcc_list_)
    print("Average eval coco_psnr:{},coco_ssim:{}".format(coco_val_psnr, coco_val_ssim))
    print("Average eval coco_mse:{},coco_npcc:{}".format(coco_val_mse, coco_val_npcc))
    # ws.cell(row=1, column=9).value = "coco_mse"
    # ws.cell(row=1, column=10).value = "coco_psnr"
    # ws.cell(row=1, column=11).value = "coco_npcc"
    # ws.cell(row=1, column=12).value = "coco_ssim"
    # ws.cell(row=s, column=9).value = coco_val_mse.item()
    # ws.cell(row=s, column=10).value = coco_val_psnr
    # ws.cell(row=s, column=11).value = coco_val_npcc.item()
    # ws.cell(row=s, column=12).value = coco_val_ssim.item()

    # 验证lfw数据集
    lfw_psnr_list_ = []
    lfw_ssim_list_ = []
    lfw_mse_list_ = []
    lfw_npcc_list_ = []
    for batch_id, val_data in enumerate(lfw_val_loader):
        with torch.no_grad():
            imgs, label = val_data
            imgs = imgs.to(device)
            label = label.to(device)
            output = RDR_model(imgs)
            # output = torch.where(output < 1, output, b)

        # Calculate  eval average psnr, ssim

        lfw_psnr_list_.append(to_psnr(output, label))
        lfw_ssim_list_.append(to_ssim(output, label))
        lfw_mse_list_.append(to_mseloss(output, label))
        lfw_npcc_list_.append(to_pearson(output, label))

    lfw_val_psnr = sum(lfw_psnr_list_) / len(lfw_psnr_list_)
    lfw_val_ssim = sum(lfw_ssim_list_) / len(lfw_ssim_list_)
    lfw_val_mse = sum(lfw_mse_list_) / len(lfw_mse_list_)
    lfw_val_npcc = sum(lfw_npcc_list_) / len(lfw_npcc_list_)
    print("Average eval lfw_psnr:{},lfw_ssim:{}".format(lfw_val_psnr, lfw_val_ssim))
    print("Average eval lfw_mse:{},lfw_npcc:{}".format(lfw_val_mse, lfw_val_npcc))
    # ws.cell(row=1, column=13).value = "lfw_mse"
    # ws.cell(row=1, column=14).value = "lfw_psnr"
    # ws.cell(row=1, column=15).value = "lfw_npcc"
    # ws.cell(row=1, column=16).value = "lfw_ssim"
    # ws.cell(row=s, column=13).value = lfw_val_mse.item()
    # ws.cell(row=s, column=14).value = lfw_val_psnr
    # ws.cell(row=s, column=15).value = lfw_val_npcc.item()
    # ws.cell(row=s, column=16).value = lfw_val_ssim.item()

    # 验证celeb数据集
    celeb_psnr_list_ = []
    celeb_ssim_list_ = []
    celeb_mse_list_ = []
    celeb_npcc_list_ = []
    for batch_id, val_data in enumerate(celeb_val_loader):
        with torch.no_grad():
            imgs, label = val_data
            imgs = imgs.to(device)
            label = label.to(device)
            output = RDR_model(imgs)
            # output = torch.where(output < 1, output, b)

        # Calculate  eval average psnr, ssim

        celeb_psnr_list_.append(to_psnr(output, label))
        celeb_ssim_list_.append(to_ssim(output, label))
        celeb_mse_list_.append(to_mseloss(output, label))
        celeb_npcc_list_.append(to_pearson(output, label))

    celeb_val_psnr = sum(celeb_psnr_list_) / len(celeb_psnr_list_)
    celeb_val_ssim = sum(celeb_ssim_list_) / len(celeb_ssim_list_)
    celeb_val_mse = sum(celeb_mse_list_) / len(celeb_mse_list_)
    celeb_val_npcc = sum(celeb_npcc_list_) / len(celeb_npcc_list_)
    print("Average eval celeb_psnr:{},celeb_ssim:{}".format(celeb_val_psnr, celeb_val_ssim))
    print("Average eval celeb_mse:{},celeb_npcc:{}".format(celeb_val_mse, celeb_val_npcc))
    # ws.cell(row=1, column=17).value = "celeb_mse"
    # ws.cell(row=1, column=18).value = "celeb_psnr"
    # ws.cell(row=1, column=19).value = "celeb_npcc"
    # ws.cell(row=1, column=20).value = "celeb_ssim"
    # ws.cell(row=s, column=17).value = celeb_val_mse.item()
    # ws.cell(row=s, column=18).value = celeb_val_psnr
    # ws.cell(row=s, column=19).value = celeb_val_npcc.item()
    # ws.cell(row=s, column=20).value = celeb_val_ssim.item()

    # 验证imagenet数据集
    img_psnr_list_ = []
    img_ssim_list_ = []
    img_mse_list_ = []
    img_npcc_list_ = []
    for batch_id, val_data in enumerate(img_val_loader):
        with torch.no_grad():
            imgs, label = val_data
            imgs = imgs.to(device)
            label = label.to(device)
            output = RDR_model(imgs)
            # output = torch.where(output < 1, output, b)

        # Calculate  eval average psnr, ssim

        img_psnr_list_.append(to_psnr(output, label))
        img_ssim_list_.append(to_ssim(output, label))
        img_mse_list_.append(to_mseloss(output, label))
        img_npcc_list_.append(to_pearson(output, label))

    img_val_psnr = sum(img_psnr_list_) / len(img_psnr_list_)
    img_val_ssim = sum(img_ssim_list_) / len(img_ssim_list_)
    img_val_mse = sum(img_mse_list_) / len(img_mse_list_)
    img_val_npcc = sum(img_npcc_list_) / len(img_npcc_list_)
    print("Average eval img_psnr:{},img_ssim:{}".format(img_val_psnr, img_val_ssim))
    print("Average eval img_mse:{},img_npcc:{}".format(img_val_mse, img_val_npcc))
    # ws.cell(row=1, column=21).value = "img_mse"
    # ws.cell(row=1, column=22).value = "img_psnr"
    # ws.cell(row=1, column=23).value = "img_npcc"
    # ws.cell(row=1, column=24).value = "img_ssim"
    # ws.cell(row=s, column=21).value = img_val_mse.item()
    # ws.cell(row=s, column=22).value = img_val_psnr
    # ws.cell(row=s, column=23).value = img_val_npcc.item()
    # ws.cell(row=s, column=24).value = img_val_ssim.item()
    # wb.save('epoch_d=f.xlsx')
    empty_cache()

    if img_val_psnr >= old_psnr:   # and val_ssim <= old_ssim:
        torch.save(RDR_model.state_dict(), 'parameter/npcc_gr36_1.5f')
        old_psnr = img_val_psnr
        # old_ssim = val_ssim/'''''''''“
        print("权重更新完成")
    # if i == 100:
    #     torch.save(RDR_model.state_dict(), 'weight_parameter_gr36{}'.format(i))

    one_epoch_time = time.time() - start_time
    print("one epoch time:{}".format(one_epoch_time))

'''
L2正则化，剔除bias and bn
weight_decay_list = (param for name, param in model.named_parameters() if name[-4:] != 'bias' and "bn" not in name)
no_decay_list = (param for name, param in model.named_parameters() if name[-4:] == 'bias' or "bn" in name)
parameters = [{'params': weight_decay_list},
              {'params': no_decay_list, 'weight_decay': 0.}]
 
optimizer = torch.optimizer.SGD(parameters, lr=0.1, momentum=0.9, weight_decay=5e-4, nesterov=True)
'''
